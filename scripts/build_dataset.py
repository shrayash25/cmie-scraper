#!/usr/bin/env python3
"""
Consolidate the downloaded CMIE report files into two shareable datasets:

  1) output/share/cmie_dataset_long.csv  — tidy "long" format, ONE ROW PER
     DATAPOINT with full provenance (industry, category, table, unit, frequency,
     indicator, period, value, count, repnum, source_url, fetched_at, sha256).
     Opens directly in Excel; rebuild any table with a PivotTable; DB-ready.

  2) output/share/cmie_dataset.xlsx — a multi-sheet workbook: an Index sheet
     plus one readable sheet per table (indicators down the side, periods across
     the top), with unit/frequency/source in the header. For human readers.

Reuses the parser + period formatter from view_report. Run after downloading:
  python scripts/build_dataset.py
"""
from __future__ import annotations

import argparse
import csv
import json

import cmie_common as cc
from view_report import fmt_period, parse_cmie_txt

LONG_CSV = cc.ROOT / "output" / "share" / "cmie_dataset_long.csv"
XLSX = cc.ROOT / "output" / "share" / "cmie_dataset.xlsx"

LONG_COLS = ["industry", "category", "table", "unit", "frequency", "indicator",
             "period", "period_raw", "value", "count", "repnum", "repcode",
             "source_url", "fetched_at", "sha256", "source_file"]


def num(v: str):
    """Cast to number when possible (for Excel); else keep text / None for blank."""
    v = (v or "").strip()
    if v == "":
        return None
    try:
        return float(v.replace(",", "")) if ("." in v or "," in v) else int(v)
    except ValueError:
        return v


def catalog_lookup() -> dict:
    lut = {}
    if cc.CATALOG.exists():
        for r in csv.DictReader(open(cc.CATALOG)):
            lut[(r["frequency"], r["repnum"])] = r
    return lut


def provenance(stem: str) -> dict:
    m = cc.RAW_DIR / f"{stem}.meta.json"
    if m.exists():
        try:
            return json.loads(m.read_text())
        except Exception:
            return {}
    return {}


def metric_indices(header):
    """Column indices that are real indicators (exclude Year, Frequency, Count)."""
    count_idx = next((j for j, h in enumerate(header) if h.strip().lower() == "count"), None)
    metrics = [j for j in range(2, len(header)) if j != count_idx]
    return metrics, count_idx


def iter_rows(files, lut):
    for f in files:
        stem = f.stem                       # e.g. A_183350
        repnum = stem.split("_", 1)[-1]
        p = parse_cmie_txt(f)
        cat = lut.get((p["freq"], repnum), {})
        prov = provenance(stem)
        metrics, count_idx = metric_indices(p["header"])
        for row in p["data"]:
            period_raw = row[0].strip() if row else ""
            count = row[count_idx].strip() if (count_idx is not None and count_idx < len(row)) else ""
            for j in metrics:
                yield {
                    "industry": p["industry"],
                    "category": cat.get("tab_label", ""),
                    "table": p["table"],
                    "unit": p["unit"],
                    "frequency": p["freq"],
                    "indicator": p["header"][j].strip(),
                    "period": fmt_period(period_raw, p["freq"]),
                    "period_raw": period_raw,
                    "value": row[j].strip() if j < len(row) else "",
                    "count": count,
                    "repnum": repnum,
                    "repcode": cat.get("repcode", ""),
                    "source_url": prov.get("request_url", ""),
                    "fetched_at": prov.get("fetched_at_utc", ""),
                    "sha256": prov.get("sha256", ""),
                    "source_file": f.name,
                }


def write_long_csv(files, lut) -> int:
    LONG_CSV.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with open(LONG_CSV, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=LONG_COLS)
        w.writeheader()
        for r in iter_rows(files, lut):
            w.writerow(r)
            n += 1
    return n


def write_xlsx(files, lut):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    idx = wb.active
    idx.title = "Index"
    idx.append(["Category", "Table", "Unit", "Freq", "Periods", "repnum", "Sheet"])
    for c in idx[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F2740")
    idx.freeze_panes = "A2"

    used = set()

    def sheet_name(title, repnum):
        import re
        base = re.sub(r"[\\/?*\[\]:]", " ", title)[:24].strip()
        name = f"{base} {repnum}"[:31]
        i = 1
        while name in used:
            name = f"{base[:22]} #{i}"[:31]
            i += 1
        used.add(name)
        return name

    for f in sorted(files):
        repnum = f.stem.split("_", 1)[-1]
        p = parse_cmie_txt(f)
        cat = lut.get((p["freq"], repnum), {})
        periods = [fmt_period(r[0], p["freq"]) for r in p["data"]]
        sn = sheet_name(p["table"], repnum)
        ws = wb.create_sheet(sn)

        ws["A1"] = p["table"]; ws["A1"].font = Font(bold=True, size=13)
        ws["A2"] = f'{p["industry"]}   ·   {p["unit"]}   ·   frequency {p["freq"]}'
        ws["A2"].font = Font(italic=True, color="6B7280")
        ws["A3"] = f'Source: {provenance(f.stem).get("request_url", "")}'
        ws["A3"].font = Font(size=9, color="9CA3AF")

        head = ["Indicator"] + periods
        ws.append([])           # row 4 spacer
        ws.append(head)         # row 5 header
        for c in ws[5]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1E3A5F")
            c.alignment = Alignment(horizontal="right")
        ws["A5"].alignment = Alignment(horizontal="left")

        for j in range(2, len(p["header"])):
            ws.append([p["header"][j].strip()] + [num(r[j] if j < len(r) else "") for r in p["data"]])

        ws.freeze_panes = "B6"
        ws.column_dimensions["A"].width = 42
        idx.append([cat.get("tab_label", ""), p["table"], p["unit"], p["freq"],
                    f"{periods[0]}–{periods[-1]}" if periods else "", repnum, sn])

    idx.column_dimensions["B"].width = 40
    idx.column_dimensions["A"].width = 22
    XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--no-xlsx", action="store_true", help="skip the Excel workbook")
    args = ap.parse_args()

    files = sorted(cc.DATA_DIR.glob("*.txt"))
    if not files:
        raise SystemExit(f"No .txt files in {cc.DATA_DIR}")
    lut = catalog_lookup()

    n = write_long_csv(files, lut)
    print(f"long CSV : {n} datapoints from {len(files)} tables -> {LONG_CSV}")
    if not args.no_xlsx:
        write_xlsx(files, lut)
        print(f"workbook : {len(files)} sheets + Index -> {XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
